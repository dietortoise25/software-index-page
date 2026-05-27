"""提取器测试：订单Excel → 清洗后DataFrame"""
import io
import pytest
import openpyxl
from extractors.shopee_orders import extract_orders


def _make_orders_xlsx(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "orders"
    headers = ["ID do pedido", "Status do pedido", "Data de criação do pedido",
               "Nome do Produto", "Valor Total", "Preço original", "Quantidade"]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    for r, row in enumerate(rows, 2):
        for c, v in enumerate(row, 1):
            ws.cell(r, c, v)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


class TestExtractOrders:
    def test_extracts_order_count(self):
        buf = _make_orders_xlsx([
            ["A001", "Concluído", "2026-05-01 12:00", "Produto X", 30.0, 30.0, 1],
            ["A002", "Cancelado", "2026-05-02 15:00", "Produto Y", 25.0, 25.0, 1],
        ])
        df = extract_orders(buf)
        assert len(df) == 2

    def test_parses_dates(self):
        buf = _make_orders_xlsx([
            ["A001", "Concluído", "2026-05-01 12:00", "Produto X", 30.0, 30.0, 1],
        ])
        df = extract_orders(buf)
        assert hasattr(df.iloc[0]['date'], 'date') or str(df.iloc[0]['date']) == '2026-05-01'

    def test_skus_are_stripped(self):
        buf = _make_orders_xlsx([
            ["A001", "Concluído", "2026-05-01 12:00", "  Produto com espacos  ", 30.0, 30.0, 1],
        ])
        df = extract_orders(buf)
        assert df.iloc[0]['sku_name'] == "Produto com espacos"

    def test_handles_missing_column(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "orders"
        ws.cell(1, 1, "ID do pedido")
        ws.cell(1, 2, "Status do pedido")
        ws.cell(2, 1, "A001")
        ws.cell(2, 2, "Concluído")
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        df = extract_orders(buf)
        assert len(df) == 1
        assert df.iloc[0]['Valor Total'] == 0.0
