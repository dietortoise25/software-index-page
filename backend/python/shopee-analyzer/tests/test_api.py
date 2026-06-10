"""测试 FastAPI 端点"""
import io
import pytest
from fastapi.testclient import TestClient
from main import app

client = TestClient(app, raise_server_exceptions=False)


def _no_stack_leak(resp):
    """断言错误响应体里不含堆栈特征（Traceback / File "... 行号）。"""
    body = resp.text
    assert "Traceback" not in body, f"响应泄露堆栈: {body[:300]}"
    assert "Traceback (most recent call last)" not in body
    assert 'File "' not in body
    assert ".py\", line" not in body

VALID_XLSX = None


def _get_valid_xlsx():
    global VALID_XLSX
    if VALID_XLSX is None:
        import openpyxl
        wb = openpyxl.Workbook()
        # 最小 8 张工作表
        sheet_keys = [
            ("已下订单", ["日期", "销售额 (BRL)", "销售额（扣除Shopee补贴）", "订单数",
                          "商品点击量", "访客数", "订单转化率", "买家数", "新买家数",
                          "现有买家数量", "潜在买家数", "已取消的订单", "已退货/退款的订单"]),
            ("已付款订单", ["日期", "销售额 (BRL)", "销售额（扣除Shopee补贴）", "订单数",
                           "商品点击量", "访客数", "订单转化率", "买家数", "新买家数",
                           "现有买家数量", "潜在买家数", "已取消的订单", "已退货/退款的订单"]),
            ("流量来源（已下订单）", ["流量来源", "销售 (BRL)", "销售占比", "商品曝光量", "商品点击量",
                                    "点击率", "订单转化率", "订单数"]),
            ("来源分布（已下订单）", ["流量来源", "日期", "商品曝光量", "商品点击量", "订单数"]),
            ("商品分布（已下订单）", ["商品编号", "商品", "销售 (BRL)", "销售占比", "商品曝光量",
                                    "商品点击量", "订单数", "点击率", "订单转化率"]),
            ("流量来源（已付款的订单）", ["流量来源", "销售 (BRL)", "销售占比", "商品曝光量", "商品点击量",
                                       "点击率", "订单转化率", "订单数"]),
            ("来源分布（已付款订单）", ["流量来源", "日期", "商品曝光量", "商品点击量", "订单数"]),
            ("商品分布（已付款订单）", ["商品编号", "商品", "销售 (BRL)", "销售占比", "商品曝光量",
                                     "商品点击量", "订单数", "点击率", "订单转化率"]),
        ]
        # 移除默认 sheet
        wb.remove(wb.active)
        for name, headers in sheet_keys:
            ws = wb.create_sheet(name)
            # 写标题行
            for c, h in enumerate(headers, 1):
                ws.cell(1, c, h)
            # 写一行日数据
            ws.cell(2, 1, "17/05/2026")
            for c in range(2, len(headers) + 1):
                ws.cell(2, c, "0")

        buf = io.BytesIO()
        wb.save(buf)
        VALID_XLSX = buf.getvalue()

    return VALID_XLSX


class TestHealth:
    def test_health(self):
        resp = client.get("/api/health")
        assert resp.status_code == 200
        assert resp.json()["status"] == "ok"


class TestAnalyzeEndpoint:
    def test_valid_xlsx(self):
        resp = client.post(
            "/api/analyze",
            files={"file": ("test.xlsx", _get_valid_xlsx(),
                   "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert resp.status_code == 200
        data = resp.json()
        assert "orders" in data
        assert "paid_orders" in data
        assert "traffic" in data
        assert "products" in data
        assert "users" in data

    def test_invalid_extension(self):
        resp = client.post(
            "/api/analyze",
            files={"file": ("test.txt", b"not excel", "text/plain")},
        )
        assert resp.status_code == 400

    def test_empty_file(self):
        resp = client.post(
            "/api/analyze",
            files={"file": ("empty.xlsx", b"", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert resp.status_code == 400

    def test_not_excel_content(self):
        resp = client.post(
            "/api/analyze",
            files={"file": ("fake.xlsx", b"not a zip file at all",
                   "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert resp.status_code == 400

    def test_missing_sheets(self):
        """Excel 缺少必需工作表时返回 422"""
        import openpyxl
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("无关数据")
        ws.cell(1, 1, "随便")
        buf = io.BytesIO()
        wb.save(buf)
        resp = client.post(
            "/api/analyze",
            files={"file": ("bad.xlsx", buf.getvalue(),
                   "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert resp.status_code in (422, 500)


class TestStoreRouterMounted:
    """店铺分析路由拆出后仍可用，端点路径/正常行为不变"""

    def test_store_router_imports(self):
        from routers.shopee_store import router as store_router
        assert store_router is not None

    def test_health(self):
        resp = client.get("/api/health")
        assert resp.status_code == 200
        assert resp.json()["status"] == "ok"

    def test_get_rules(self):
        resp = client.get("/api/rules")
        assert resp.status_code == 200
        assert "rules" in resp.json()

    def test_simulate_ok(self):
        resp = client.post("/api/simulate", json={"ad_metrics": {}, "realloc_pct": 70})
        assert resp.status_code == 200
        assert "baseline" in resp.json()


class TestNoStackLeak:
    """错误响应不得包含堆栈，且返回有意义的 message"""

    def test_analyze_error_no_stack(self, monkeypatch):
        import routers.shopee_store as store_mod

        def boom(*a, **k):
            raise RuntimeError("内部解析炸了")

        monkeypatch.setattr(store_mod, "analyze_excel", boom)
        # 通过前置校验（PK 头 + 扩展名），再在 analyze_excel 处抛错
        resp = client.post(
            "/api/analyze",
            files={"file": ("real.xlsx", b"PK" + b"x" * 50,
                   "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert resp.status_code == 500
        _no_stack_leak(resp)
        body = resp.json()
        detail = body.get("detail", body)
        assert "message" in detail
        assert detail["message"]

    def test_simulate_error_no_stack(self, monkeypatch):
        import routers.shopee_store as store_mod

        def boom(*a, **k):
            raise RuntimeError("模拟炸了")

        monkeypatch.setattr(store_mod, "run_simulation", boom)
        resp = client.post("/api/simulate", json={"ad_metrics": {}, "realloc_pct": 70})
        assert resp.status_code == 500
        _no_stack_leak(resp)
        detail = resp.json().get("detail", resp.json())
        assert "message" in detail and detail["message"]

    def test_search_image_error_no_stack(self, monkeypatch):
        import routers.shopee_store as store_mod

        def boom(*a, **k):
            raise RuntimeError("图搜炸了")

        monkeypatch.setattr(store_mod, "search_by_image", boom)
        resp = client.post("/api/search-image", json={"image_url": "http://x/a.jpg"})
        assert resp.status_code == 500
        _no_stack_leak(resp)
        detail = resp.json().get("detail", resp.json())
        assert "message" in detail and detail["message"]

    def test_get_rules_error_no_stack(self, monkeypatch):
        import routers.shopee_store as store_mod

        def boom(*a, **k):
            raise RuntimeError("规则读取炸了")

        monkeypatch.setattr(store_mod, "load_rules", boom)
        resp = client.get("/api/rules")
        assert resp.status_code == 500
        _no_stack_leak(resp)
        detail = resp.json().get("detail", resp.json())
        assert "message" in detail and detail["message"]
