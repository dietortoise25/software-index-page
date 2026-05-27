"""Shopee 订单 Excel 提取器"""
import pandas as pd
import openpyxl
from io import BytesIO


def extract_orders(file_source: BytesIO) -> pd.DataFrame:
    wb = openpyxl.load_workbook(file_source, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [str(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {}
        for c, h in enumerate(headers, 1):
            row[h] = ws.cell(r, c).value
        rows.append(row)
    df = pd.DataFrame(rows)
    if df.empty:
        return df
    if 'Valor Total' in df.columns:
        df['Valor Total'] = pd.to_numeric(df['Valor Total'], errors='coerce').fillna(0)
    else:
        df['Valor Total'] = 0.0
    if 'Nome do Produto' in df.columns:
        df['sku_name'] = df['Nome do Produto'].str.strip()
    if 'Data de criação do pedido' in df.columns:
        df['date'] = pd.to_datetime(df['Data de criação do pedido'], format='%Y-%m-%d %H:%M', errors='coerce')
    return df
